import streamlit as st
from groq import Groq
import json
import re
import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="🧪 Test Case Generator",
    page_icon="🧪",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
.metric-card {
    background: #f8f9fa;
    border-radius: 10px;
    padding: 16px 20px;
    text-align: center;
    border: 1px solid #e9ecef;
}
.metric-val { font-size: 32px; font-weight: 700; margin: 0; }
.metric-lbl { font-size: 13px; color: #6c757d; margin: 0; }
</style>
""", unsafe_allow_html=True)

# ── API KEY ───────────────────────────────────────────────────────────────────
api_key = ""
if hasattr(st, "secrets"):
    api_key = st.secrets.get("GROQ_API_KEY", "")

with st.sidebar:
    st.title("🧪 Test Case Generator")
    st.caption("AI-генерация тест-кейсов из ТЗ")
    st.divider()

    if not api_key:
        api_key = st.text_input(
            "🔑 Groq API Key",
            type="password",
            placeholder="gsk_...",
            help="Получить бесплатно: https://console.groq.com",
        )

    if api_key:
        st.success("✅ API Key подключён")
    else:
        st.warning("Введи Anthropic API Key")

    st.divider()
    st.subheader("⚙️ Настройки")
    n_cases = st.slider("Количество тест-кейсов", 5, 30, 12)

    st.markdown("**Типы сценариев**")
    col_a, col_b, col_c = st.columns(3)
    with col_a: inc_pos = st.checkbox("✅ Pos", value=True)
    with col_b: inc_neg = st.checkbox("❌ Neg", value=True)
    with col_c: inc_bnd = st.checkbox("⚠️ Bnd", value=True)

    st.markdown("**Приоритеты**")
    col_d, col_e, col_f = st.columns(3)
    with col_d: inc_hi  = st.checkbox("🔴 High",  value=True)
    with col_e: inc_med = st.checkbox("🟡 Med",   value=True)
    with col_f: inc_lo  = st.checkbox("🟢 Low",   value=True)

    lang = st.selectbox("Язык тест-кейсов", ["Русский", "English"])
    st.divider()
    st.caption("Powered by Llama 3.3 · Groq")

if not api_key:
    st.title("🧪 Test Case Generator из ТЗ")
    st.info("👈 Введи GROG API Key в боковой панели чтобы начать")
    st.stop()

client = Groq(api_key=api_key)

# ── SYSTEM PROMPT строится внутри функции ─────────────────────────────────────
types_needed = []
if inc_pos: types_needed.append("positive")
if inc_neg: types_needed.append("negative")
if inc_bnd: types_needed.append("boundary")
types_str = ", ".join(types_needed) if types_needed else "positive, negative, boundary"
lang_str = "Russian" if lang == "Русский" else "English"

# ── HELPERS ───────────────────────────────────────────────────────────────────
def to_str(v):
    if isinstance(v, list): return ", ".join(str(x) for x in v)
    return str(v) if v is not None else ""

def generate_test_cases(tz_text: str) -> list[dict]:
    if len(tz_text) > 8000:
        tz_text = tz_text[:8000] + "\n...[truncated]"

    system_prompt = (
        f"You are a senior QA engineer. Generate exactly {n_cases} test cases as a JSON array. "
        "Return ONLY the JSON array, no markdown, no explanation. "
        'Each item: {"id":"TC-001","title":"...","type":"positive|negative|boundary",'
        '"priority":"high|medium|low","preconditions":"string","steps":["..."],'
        '"expected_result":"string","tags":["..."]}. '
        f"Cover types: {types_str}. Write in {lang_str}."
    )

    try:
        response = client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            max_tokens=4096,
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": f"Requirements:\n\n{tz_text}"},
            ],
            temperature=0.3,
        )
    except Exception as e:
        st.error(f"API Error: {e}")
        return []

    raw = response.choices[0].message.content.strip()
    raw = re.sub(r"```(?:json)?", "", raw).strip().rstrip("`").strip()
    m = re.search(r"\[.*\]", raw, re.DOTALL)
    if m:
        try: return json.loads(m.group())
        except: pass
    try: return json.loads(raw)
    except Exception as e:
        st.error(f"Ошибка парсинга JSON: {e}")
        with st.expander("Сырой ответ"): st.code(raw[:2000])
        return []

def to_excel(tcs: list[dict]) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    headers = ["ID","Title","Type","Priority","Preconditions","Steps","Expected Result","Tags"]
    hf = PatternFill("solid", fgColor="1F4E79")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hf
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    colors = {"positive":"E8F5E9","negative":"FFEBEE","boundary":"FFF8E1"}
    for row, tc in enumerate(tcs, 2):
        fill = PatternFill("solid", fgColor=colors.get(tc.get("type",""),"FFFFFF"))
        steps_raw = tc.get("steps", [])
        if isinstance(steps_raw, str): steps_raw = [steps_raw]
        steps_text = "\n".join(f"{i+1}. {s}" for i, s in enumerate(steps_raw))
        vals = [to_str(tc.get("id")), to_str(tc.get("title")), to_str(tc.get("type")),
                to_str(tc.get("priority")), to_str(tc.get("preconditions")),
                steps_text, to_str(tc.get("expected_result")),
                ", ".join(str(t) for t in (tc.get("tags") or []))]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for col, w in zip(range(1,9), [10,35,12,10,30,50,35,20]):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    ws2 = wb.create_sheet("Stats")
    for r, (lbl, val) in enumerate([
        ("Total",    len(tcs)),
        ("Positive", sum(1 for t in tcs if t.get("type")=="positive")),
        ("Negative", sum(1 for t in tcs if t.get("type")=="negative")),
        ("Boundary", sum(1 for t in tcs if t.get("type")=="boundary")),
        ("",""),
        ("High",   sum(1 for t in tcs if t.get("priority")=="high")),
        ("Medium", sum(1 for t in tcs if t.get("priority")=="medium")),
        ("Low",    sum(1 for t in tcs if t.get("priority")=="low")),
    ], 1):
        ws2.cell(row=r, column=1, value=lbl).font = Font(bold=True)
        ws2.cell(row=r, column=2, value=val)
    ws2.column_dimensions["A"].width = 18
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ── MAIN ──────────────────────────────────────────────────────────────────────
st.title("🧪 Test Case Generator из ТЗ")
st.markdown("Загрузи техническое задание — AI сгенерирует тест-кейсы с экспортом в Excel и JSON")

tab1, tab2 = st.tabs(["✏️ Текст", "📁 Файл (PDF / DOCX)"])
tz_text = ""

with tab1:
    inp = st.text_area("Текст ТЗ:", height=280,
                        placeholder="Вставьте текст технического задания...")
    if inp:
        tz_text = inp
        st.caption(f"{len(tz_text):,} / 8,000 символов")

with tab2:
    uploaded = st.file_uploader("PDF или DOCX", type=["pdf","docx","doc"])
    if uploaded:
        import tempfile
        from pathlib import Path
        suf = Path(uploaded.name).suffix.lower()
        with tempfile.NamedTemporaryFile(delete=False, suffix=suf) as tmp:
            tmp.write(uploaded.read()); tmp_path = tmp.name
        try:
            if suf == ".pdf":
                import fitz
                doc = fitz.open(tmp_path)
                tz_text = "\n\n".join(p.get_text() for p in doc)
                doc.close()
            elif suf in (".docx", ".doc"):
                from docx import Document
                doc = Document(tmp_path)
                tz_text = "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())
            st.success(f"✅ {uploaded.name} — {len(tz_text):,} символов")
            with st.expander("Просмотр"):
                st.text(tz_text[:2000] + ("..." if len(tz_text) > 2000 else ""))
        except Exception as e:
            st.error(f"Ошибка: {e}")

st.divider()
if st.button("🚀 Генерировать тест-кейсы", type="primary",
             use_container_width=True, disabled=not bool(tz_text)):
    with st.spinner("Генерируем..."):
        tcs = generate_test_cases(tz_text)
        if tcs:
            st.session_state["tcs"] = tcs
            st.success(f"✅ Сгенерировано {len(tcs)} тест-кейсов")
        else:
            st.error("Не удалось сгенерировать. Проверь API ключ.")

if "tcs" in st.session_state and st.session_state["tcs"]:
    all_tcs = st.session_state["tcs"]
    active_types = [t for t,on in [("positive",inc_pos),("negative",inc_neg),("boundary",inc_bnd)] if on]
    active_prios = [p for p,on in [("high",inc_hi),("medium",inc_med),("low",inc_lo)] if on]
    tcs = [tc for tc in all_tcs if tc.get("type") in active_types and tc.get("priority") in active_prios]

    st.divider()
    c1,c2,c3,c4 = st.columns(4)
    with c1: st.markdown(f'<div class="metric-card"><p class="metric-val">{len(all_tcs)}</p><p class="metric-lbl">Всего</p></div>', unsafe_allow_html=True)
    with c2: st.markdown(f'<div class="metric-card"><p class="metric-val" style="color:#28a745">{sum(1 for t in all_tcs if t.get("type")=="positive")}</p><p class="metric-lbl">✅ Positive</p></div>', unsafe_allow_html=True)
    with c3: st.markdown(f'<div class="metric-card"><p class="metric-val" style="color:#dc3545">{sum(1 for t in all_tcs if t.get("type")=="negative")}</p><p class="metric-lbl">❌ Negative</p></div>', unsafe_allow_html=True)
    with c4: st.markdown(f'<div class="metric-card"><p class="metric-val" style="color:#fd7e14">{sum(1 for t in all_tcs if t.get("type")=="boundary")}</p><p class="metric-lbl">⚠️ Boundary</p></div>', unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    col1, col2 = st.columns(2)
    with col1:
        st.download_button("📥 Excel (Zephyr/Jira)", to_excel(tcs), "test_cases.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)
    with col2:
        st.download_button("📥 JSON (Zephyr API)", json.dumps(tcs, ensure_ascii=False, indent=2),
                           "test_cases.json", "application/json", use_container_width=True)

    st.subheader(f"📋 Тест-кейсы ({len(tcs)} из {len(all_tcs)})")
    ti = {"positive":"✅","negative":"❌","boundary":"⚠️"}
    pi = {"high":"🔴","medium":"🟡","low":"🟢"}
    for tc in tcs:
        label = f"{ti.get(tc.get('type',''),'?')} **{tc.get('id','?')}** — {tc.get('title','?')}  {pi.get(tc.get('priority','medium'),'⚪')}"
        with st.expander(label):
            l, r = st.columns([1,2])
            with l:
                st.markdown(f"**Тип:** `{tc.get('type')}`")
                st.markdown(f"**Приоритет:** `{tc.get('priority')}`")
                st.markdown(f"**Предусловия:**  \n{to_str(tc.get('preconditions','—'))}")
                if tc.get("tags"):
                    st.markdown("**Теги:** " + " ".join(f"`{t}`" for t in tc["tags"]))
            with r:
                st.markdown("**Шаги:**")
                steps = tc.get("steps",[])
                if isinstance(steps, str): steps = [steps]
                for idx, step in enumerate(steps, 1):
                    st.markdown(f"{idx}. {step}")
                st.success(f"**Ожидаемый результат:**  \n{to_str(tc.get('expected_result','—'))}")

elif tz_text:
    st.info("👆 Нажми «Генерировать тест-кейсы»")
else:
    st.info("👆 Введи текст ТЗ или загрузи файл")